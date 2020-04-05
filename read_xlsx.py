from openpyxl import load_workbook
wb = load_workbook(filename = 'sample-file.xlsx')

# list worksheet names in XLSX workbook
print("Worksheets: ",wb.sheetnames)

# activate worksheet by worksheet name
workSheetNameToActivate = "Sample SSN numbers"
#workSheetNameToActivate = "Sheet2"
if workSheetNameToActivate in wb.sheetnames:
    activeWorkSheet = wb[workSheetNameToActivate]

try:
    if activeWorkSheet:
        # iterate for rows and cells
        # for rows in activeWorkSheet:
        #     for cell in rows:
        #         print(cell.value)

        # get column headers, i.e. only the first row
        headers = ()
        for row in activeWorkSheet.iter_rows(min_row=1, max_col=3, max_row=1, values_only=True):
            #print(row)
            headers = row

        print("\nHeaders: ",headers, "\n")

        # get values
        formattedData = []
        maxColumns = 3 # change it with columns you want to iterate
        for row in activeWorkSheet.iter_rows(min_row=2, max_col=maxColumns, max_row=3, values_only=True):
            print("Row data: ",row)
            tempData = {}
            for i in range(maxColumns):
                tempData[headers[i]] = row[i]
            formattedData.append(tempData)
            
        print("\nheader-data-pair: ",formattedData)
except Exception as e:
    print(e)


